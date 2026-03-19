import openpyxl, json, os

script_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(script_dir, "article-upload-template.xlsx")
json_path = os.path.join(script_dir, "prisma", "excel-data.json")

wb = openpyxl.load_workbook(excel_path)
ws = wb["Articles"]
headers = [cell.value for cell in ws[1]]
articles = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    art = {}
    for i, h in enumerate(headers):
        val = row[i] if i < len(row) else None
        art[h] = str(val) if val else ""
    articles.append(art)

with open(json_path, "w") as f:
    json.dump(articles, f, indent=2)

print(f"   Found {len(articles)} articles")
