from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = 'Articles'

headers = ['category', 'subCategory', 'shortTitle', 'title', 'authorName', 'sourceUrl', 'publishedAt', 'fullContent', 'mediumSummary', 'shortSummary', 'imageUrl', 'isPublished']
required = {0,1,3,6,7,8,9}
blue_fill = PatternFill('solid', fgColor='4472C4')
gray_fill = PatternFill('solid', fgColor='A5A5A5')
white_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for i, h in enumerate(headers):
    cell = ws.cell(row=1, column=i+1, value=h)
    cell.fill = blue_fill if i in required else gray_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

widths = [35, 42, 18, 65, 30, 45, 14, 80, 60, 50, 5, 12]
for i, w in enumerate(widths):
    ws.column_dimensions[chr(65+i)].width = w

data_font = Font(name='Arial', size=10)
wrap_align = Alignment(wrap_text=True, vertical='top')
top_align = Alignment(vertical='top')

articles = [
    ["Foundational Concepts & Definition", "Definition of Health AI", "AI Definitions", "Transformative Potential of AI in Healthcare: Definitions, Applications, and Navigating the Ethical Landscape and Public Perspectives", "Bekbolatova M, Mayer J, Ong CW, Toma M", "https://doi.org/10.3390/healthcare12020125", "2024-01-05",
     "This review examines artificial intelligence (AI) as a crucial tool in healthcare, aimed at improving patient outcomes and optimizing healthcare delivery. By harnessing machine learning algorithms, natural language processing, and computer vision, AI enables analysis of complex medical data. The integration of AI into healthcare systems supports clinicians, personalizes patient care, and enhances population health while addressing rising costs and limited resources. As a subdivision of computer science, AI focuses on developing advanced algorithms capable of performing complex tasks once reliant on human intelligence, with the goal of achieving human-level performance with improved efficiency and accuracy. Various industries including engineering, finance, and education have benefited from AI, and healthcare has witnessed rapid growth in AI utilization. The review emphasizes that AI should not be viewed as a threat to human workers but rather as a tool to augment and support healthcare professionals, freeing them to focus on complex and critical tasks. However, legal and ethical challenges must be addressed alongside comprehensive public education to ensure widespread acceptance.",
     "This review defines AI in healthcare as a tool for improving patient outcomes through machine learning, NLP, and computer vision. It clarifies that AI augments rather than replaces healthcare professionals, automating routine tasks so clinicians can focus on complex patient care. The paper addresses ethical and legal challenges that must be resolved for widespread AI adoption in medicine.",
     "A comprehensive review defining AI's role in healthcare, emphasizing it augments clinicians rather than replacing them, while addressing ethical and legal challenges for adoption.", "", "TRUE"],

    ["Foundational Concepts & Definition", "Definition of Health AI", "AI in Medicine", "Artificial Intelligence, the new scalpel of medicine. Part I: Definitions, key concepts, and historical overview", "Kolh P, Bellavia S", "https://pubmed.ncbi.nlm.nih.gov/41069296/", "2025-10-01",
     "The integration of artificial intelligence (AI) into healthcare is not simply an evolution but an ongoing revolution that is fundamentally redefining medical practice. Medical AI is much more than a tool; it is a new frontier in contemporary medicine. As a staple in hospitals and doctors' offices worldwide, AI systems now assist healthcare providers by analyzing massive amounts of medical data and offering personalized diagnoses and treatment recommendations. This article reviews key definitions and concepts related to AI and describes the different stages of AI development by placing them in a historical context. It covers foundational terminology including machine learning, deep learning, neural networks, and natural language processing, establishing a common language for clinicians entering the field.",
     "This article provides foundational definitions of AI in medicine, covering key concepts like machine learning, deep learning, and neural networks. It positions AI not merely as a tool but as a new frontier redefining medical practice, tracing the stages of AI development within their historical context.",
     "A foundational review of AI definitions and key concepts in medicine, covering machine learning, deep learning, and neural networks in their historical context.", "", "TRUE"],

    ["Foundational Concepts & Definition", "Definition of Health AI", "Clinical AI Support", "AI-Driven Clinical Decision Support Systems: An Ongoing Pursuit of Potential", "Elhaddad M, Hamam S", "https://doi.org/10.7759/cureus.57728", "2024-04-06",
     "Clinical Decision Support Systems (CDSS) are essential tools in contemporary healthcare, enhancing clinicians' decisions and patient outcomes. The integration of artificial intelligence (AI) is revolutionizing CDSS further. This review delves into AI technologies transforming CDSS, their applications in healthcare decision-making, associated challenges, and the potential trajectory toward fully realizing AI-CDSS potential. It examines the integration of AI technologies including machine learning algorithms like neural networks and decision trees, natural language processing, and deep learning. The review addresses challenges of AI integration such as interpretability and bias, then presents real-life examples of AI-driven diagnostics, personalized treatment recommendations, risk prediction, early intervention, and AI-assisted clinical documentation. It emphasizes user-centered design addressing usability, trust, workflow, and ethical and legal considerations.",
     "This review defines and examines AI-driven Clinical Decision Support Systems (CDSS), covering machine learning, NLP, and deep learning integration. It provides real-world examples of AI diagnostics, personalized treatment, and risk prediction while addressing challenges of interpretability, bias, and user-centered design.",
     "A review of AI-driven Clinical Decision Support Systems covering ML, NLP, and deep learning applications in diagnostics, treatment, and risk prediction.", "", "TRUE"],

    ["Foundational Concepts & Definition", "Scope and taxonomy of AI in healthcare", "AI-ECG Scope", "Artificial Intelligence Interpretation of the Electrocardiogram: A State-of-the-Art Review", "Ose B, Sattar Z, Gupta A, Toquica C, Harvey C, Noheria A", "https://doi.org/10.1007/s11886-024-02062-1", "2024-05-16",
     "Artificial intelligence (AI) is transforming electrocardiography (ECG) interpretation. AI diagnostics can reach beyond human capabilities, facilitate automated access to nuanced ECG interpretation, and expand cardiovascular screening. AI can be applied to standard 12-lead resting ECGs and single-lead ECGs in external monitors, implantable devices, and direct-to-consumer smart devices. Rhythm classification was the first application. Subsequently, models have been developed for screening structural heart disease including hypertrophic cardiomyopathy, cardiac amyloidosis, aortic stenosis, pulmonary hypertension, and left ventricular systolic dysfunction. AI models can predict future events like systolic heart failure and atrial fibrillation. AI-ECG exhibits potential in acute cardiac events and non-cardiac applications including pulmonary embolism, electrolyte abnormalities, drug monitoring, sleep apnea, and predicting all-cause mortality. Many models have received FDA clearance or breakthrough device designation.",
     "This review maps the scope and taxonomy of AI applications in ECG interpretation, from rhythm classification to structural heart disease screening, event prediction, and non-cardiac applications. It documents FDA-cleared and breakthrough-designated AI-ECG models, illustrating the breadth of AI's diagnostic reach.",
     "A comprehensive taxonomy of AI-ECG applications spanning rhythm classification, structural heart disease screening, event prediction, and FDA-cleared models.", "", "TRUE"],

    ["Foundational Concepts & Definition", "Scope and taxonomy of AI in healthcare", "Pharmacy AI Scope", "Applications of artificial intelligence in current pharmacy practice: A scoping review", "Hatzimanolis J, Riley B, El-Den S, Aslani P, Zhou J, Chaar BB", "https://doi.org/10.1016/j.sapharm.2024.12.007", "2024-12-17",
     "Artificial intelligence (AI) has been of growing research interest since its introduction to healthcare in the 1970s. Application in pharmacy practice is reportedly narrower in scope, with greater emphasis on stock management and function optimization than enhancing patient outcomes. This scoping review identified 560 studies, of which seven met inclusion criteria. All seven utilized machine learning techniques. AI identification of prescriptions requiring pharmacist intervention was most frequent (n=4), followed by screening services (n=2) and patient-facing mobile applications (n=1). Results indicated a workflow- and productivity-focused application of AI within pharmacy. The review identified three main focus areas: identification of atypical medication orders, improving efficiency of mass screening, and improving adherence and quality use of medicines. It also identified gaps in AI's current utility and its potential for day-to-day practice.",
     "This scoping review maps the scope of AI applications in pharmacy practice, identifying three main areas: medication order screening, mass screening efficiency, and medication adherence. It reveals AI in pharmacy is primarily workflow-focused rather than patient-outcome driven, highlighting gaps and untapped potential.",
     "Scoping review mapping AI's scope in pharmacy practice across medication screening, workflow optimization, and adherence improvement.", "", "TRUE"],

    ["Foundational Concepts & Definition", "Scope and taxonomy of AI in healthcare", "Transformer Taxonomy", "Task-Specific Transformer-Based Language Models in Health Care: Scoping Review", "Cho HN, Jun TJ, Kim YH, Kang H, et al.", "https://doi.org/10.2196/49724", "2024-11-18",
     "Transformer-based language models have shown great potential to revolutionize healthcare by advancing clinical decision support, patient interaction, and disease prediction. This scoping review categorizes medical transformer-based language models into six tasks: dialogue generation, question answering, summarization, text classification, sentiment analysis, and named entity recognition. Key findings revealed both advancements and critical challenges: models like MedPIR for dialogue generation show promise but face privacy concerns; question-answering models like BioBERT improve accuracy but struggle with medical terminology complexity; the BioBERTSum summarization model aids clinicians but needs better handling of long sequences. The review provides a consolidated taxonomy of transformer-based language models' roles in healthcare.",
     "This scoping review creates a taxonomy of transformer-based language models in healthcare, categorizing applications into six tasks: dialogue generation, question answering, summarization, text classification, sentiment analysis, and named entity recognition. It maps specific models to tasks and identifies challenges.",
     "A taxonomy of transformer-based language models in healthcare across six task categories from dialogue generation to named entity recognition.", "", "TRUE"],

    ["Foundational Concepts & Definition", "Scope and taxonomy of AI in healthcare", "AI in CML", "Application of artificial intelligence in chronic myeloid leukemia disease prediction and management: a scoping review", "Ram M, Afrash MR, Moulaei K, Parvin M, et al.", "https://doi.org/10.1186/s12885-024-12764-y", "2024-08-20",
     "This scoping review examines AI applications in CML management, categorizing them into tumor diagnosis/classification (75%), prediction/prognosis (17%), and treatment (8%). For tumor diagnosis, AI approaches are classified into blood smear image-based (n=5), clinical parameter-based (n=2), and gene profiling-based (n=2). The most commonly employed AI models include Support Vector Machine (SVM) (n=5), XGBoost (n=4), and various neural networks including ANN (n=3). Hybrid CNN with Interactive Autodidactic School achieved 100% accuracy in organizing leukemia data types, while MayGAN attained 99.8% accuracy in diagnosing CML from blood smear images. The review demonstrates how AI classifications provide a structured taxonomy for technology deployment across different clinical functions.",
     "This scoping review creates a taxonomy of AI applications in chronic myeloid leukemia, classifying them into diagnosis (75%), prognosis (17%), and treatment (8%). It maps AI models (SVM, XGBoost, neural networks) to specific clinical tasks with diagnostic accuracies reaching 99.8-100%.",
     "A taxonomic scoping review of AI in CML management classifying applications into diagnosis, prognosis, and treatment with models achieving up to 100% accuracy.", "", "TRUE"],

    ["Foundational Concepts & Definition", "History and evolution", "AI Evolution", "Machine Learning Tools for Acute Respiratory Distress Syndrome Detection and Prediction", "Rubulotta F, Bahrami S, Marshall DC, Komorowski M", "https://doi.org/10.1097/CCM.0000000000006390", "2024-08-12",
     "Machine learning (ML) tools for acute respiratory distress syndrome (ARDS) detection and prediction are increasingly used. ARDS is a complex lung condition arising from pneumonia, sepsis, or trauma, leading to widespread lung inflammation. ML has shown promising potential in supporting ARDS recognition in ICU patients by analyzing clinical data including vital signs, laboratory results, and imaging findings. This concise definitive review traces the evolution of AI and ML tools from early statistical methods to modern deep learning approaches for ARDS prediction and detection in critically ill patients. It chronicles how the field has progressed from simple logistic regression models to ensemble methods and neural networks, reflecting the broader evolution of AI in medicine. Each generation of AI tools has built upon previous capabilities, offering increasingly sophisticated pattern recognition for timely interventions.",
     "This review traces the evolution of ML tools for ARDS detection, chronicling progression from early logistic regression to modern deep learning. It demonstrates how AI capabilities in critical care have evolved over time, with each generation building upon previous methods for improved pattern recognition.",
     "A review tracing the evolution of ML tools from simple statistical models to modern deep learning for ARDS detection in critical care.", "", "TRUE"],

    ["Foundational Concepts & Definition", "History and evolution", "AI in PCOS", "Artificial intelligence in polycystic ovarian syndrome management: past, present, and future", "Wang J, Chen R, Long H, He J, Tang M, et al.", "https://doi.org/10.1007/s11547-025-02032-9", "2025-06-23",
     "This systematic review traces the historical evolution of AI-driven interventions in polycystic ovary syndrome (PCOS) management from early prediction models to contemporary deep learning approaches. AI-based analytics have profoundly transformed PCOS management in prediction, diagnosis, classification, and screening of complications. The review delineates a comprehensive inventory of AI applications across diverse clinical contexts, tracing progression from basic machine learning classifiers to sophisticated digital health ecosystems. It examines the potential of amalgamating existing digital health technologies to forge an AI-augmented healthcare ecosystem for prevention and holistic PCOS management, discussing strategic avenues for clinical translation.",
     "This systematic review traces AI's evolution in PCOS management from early prediction models to modern digital health ecosystems. It maps the historical progression across prediction, diagnosis, classification, and screening, demonstrating how AI capabilities have expanded from simple classifiers to comprehensive management platforms.",
     "A systematic review tracing AI's historical evolution in PCOS management from early prediction models to modern AI-augmented digital health ecosystems.", "", "TRUE"],

    ["Foundational Concepts & Definition", "History and evolution", "AI in Orthopaedics", "Artificial Intelligence in Orthopaedics: Clinical Performance, Limitations, and Translational Readiness", "Glinkowski WM, Spalinska A, Wolk A, Wolk K", "https://doi.org/10.3390/jcm15051751", "2026-02-25",
     "This narrative review synthesizes current evidence on AI in orthopaedics across five areas: diagnostic imaging, surgical planning, predictive analytics, rehabilitation intelligence, and system-level management. Covering 2019-2025, it traces AI's evolution from early fracture detection algorithms to sophisticated surgical planning systems. The most mature evidence exists for AI-assisted bone fracture detection on radiographs (sensitivity ~90%, specificity ~92%) and implant identification (97-99% accuracy). AI-based planning increases likelihood of reducing intraoperative corrections and improving functional outcomes. The review traces historical milestones showing progression from pattern recognition to integrated, patient-centered decision support throughout perioperative and rehabilitation periods.",
     "This review traces AI's evolution in orthopaedics from 2019-2025, covering imaging, surgical planning, predictive analytics, rehabilitation, and system-level applications. It documents progression from basic fracture detection (90% sensitivity) to integrated decision support, highlighting achievements and translational gaps.",
     "A review tracing AI's evolution in orthopaedics from fracture detection to integrated perioperative decision support across five clinical domains.", "", "TRUE"],

    ["Foundational Concepts & Definition", "Conceptual frameworks", "LLM Framework", "The Role of Large Language Models in Transforming Emergency Medicine: Scoping Review", "Preiksaitis C, Ashenburg N, Bunney G, Chu A, et al.", "https://doi.org/10.2196/53787", "2024-05-10",
     "Given the absence of a comprehensive framework for LLMs in emergency medicine, this scoping review systematically maps existing literature. Using PRISMA-ScR criteria, 1994 titles were screened and 43 papers included. The review uncovered four major thematic frameworks: (1) clinical decision-making and support, with LLMs enhancing patient care through real-time triage; (2) efficiency, workflow, and information management through automation of patient record synthesis; (3) risks, ethics, and transparency regarding reliability and biased decision-making from flawed training data; and (4) education and communication possibilities including simulated patient interactions. This four-pillar framework provides a conceptual structure for understanding and implementing LLMs across emergency care settings.",
     "This scoping review proposes a four-pillar conceptual framework for LLMs in emergency medicine: clinical decision support, workflow efficiency, ethics/transparency, and education/communication. Based on 43 studies, it provides a structured approach for understanding and implementing AI across emergency care.",
     "A four-pillar conceptual framework for LLMs in emergency medicine covering clinical decision support, workflow, ethics, and education.", "", "TRUE"],

    ["Foundational Concepts & Definition", "Conceptual frameworks", "Frailty AI Framework", "Use of Conventional Artificial Intelligence Methods in the Identification of Frailty: A Scoping Review", "Dalsania KA, Menard A, Sundararaman S, et al.", "https://doi.org/10.1111/jgs.70387", "2026-03-17",
     "This scoping review examines clinical tools and conceptual frameworks used as reference standards in training AI algorithms for frailty identification. Following the Arksey and O'Malley framework, eight academic databases were searched. The review included 33 publications with 23 different AI techniques identified, with logistic regression and decision trees being most common. Among 21 distinct reference standards, the Physical Frailty Phenotype was cited most frequently (n=7). Most AI methods (n=27) prioritized identification, one addressed management, and five focused on both. None engaged knowledge users in defining or validating tools, and only three explored algorithmic biases. The review reveals that emerging AI tools lack a consistent conceptual framework, leading to design and implementation inconsistencies.",
     "This scoping review examines conceptual frameworks for AI-based frailty identification, finding 23 AI techniques and 21 different reference standards across 33 studies. It reveals a critical lack of consistent conceptual frameworks, with no studies engaging end users in AI tool design.",
     "A scoping review revealing the lack of consistent conceptual frameworks in AI-based frailty identification across 33 studies using 23 different AI techniques.", "", "TRUE"],

    ["Foundational Concepts & Definition", "Conceptual frameworks", "Metaverse Framework", "Immersive futures in healthcare: A mapping review of review articles on the metaverse", "Rikala P, Ylonen M, Solberg M, Sellberg C, et al.", "https://doi.org/10.1177/20552076261431602", "2026-03-13",
     "This mapping review synthesizes existing review studies on the metaverse in healthcare using topic modeling, hierarchical clustering, and qualitative interpretation. The analysis yielded three distinct conceptual clusters: (1) immersive therapeutic and educational applications with intelligent integration, (2) immersive technologies for surgical training and clinical simulation, and (3) integrated, immersive, and intelligent technologies for personalized, networked healthcare. These clusters illustrate a shift from conceptual exploration toward applied, system-level integration. Applications show promise in mental health, surgical education, and personalized care, but evidence is preliminary. Key risks include privacy concerns, governance gaps, and equity challenges. The review aligns with WHO strategic objectives for responsible adoption.",
     "This mapping review proposes a three-cluster conceptual framework for metaverse technologies in healthcare: therapeutic/educational applications, surgical training/simulation, and personalized networked healthcare. It aligns with WHO strategic objectives and identifies risks around privacy and equity.",
     "A three-cluster conceptual framework for metaverse technologies in healthcare covering therapeutic, surgical, and personalized care applications.", "", "TRUE"],
]

for r, row_data in enumerate(articles, 2):
    for i, val in enumerate(row_data):
        cell = ws.cell(row=r, column=i+1, value=val)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = wrap_align if i >= 7 else top_align
    ws.row_dimensions[r].height = 100

ws.auto_filter.ref = 'A1:L14'
ws.freeze_panes = 'A2'

ins = wb.create_sheet('Instructions')
ins.sheet_properties.tabColor = '4472C4'
ins.column_dimensions['A'].width = 90
title_font = Font(name='Arial', bold=True, size=16, color='1F4E79')
section_font = Font(name='Arial', bold=True, size=12, color='2E75B6')
body_font = Font(name='Arial', size=10)
note_font = Font(name='Arial', size=10, italic=True, color='666666')

instr_rows = [
    ('Health AI Digest - Article Upload Template', title_font),
    ('', None),
    ('REQUIRED FIELDS (Blue Headers):', section_font),
    ('  category - Level 1 main category (appears as a card on the homepage)', body_font),
    ('  subCategory - Level 2 sub-theme (appears as a card within the category page)', body_font),
    ('  title - Full article title', body_font),
    ('  publishedAt - Publication date in YYYY-MM-DD format', body_font),
    ('  fullContent - Detailed Summary (recommended: 400-800 words)', body_font),
    ('  mediumSummary - Brief Summary (recommended: 100-250 words)', body_font),
    ('  shortSummary - Quick Read (recommended: 30-60 words)', body_font),
    ('', None),
    ('OPTIONAL FIELDS (Gray Headers):', section_font),
    ('  shortTitle - 2-3 word label displayed above the article title on cards', body_font),
    ('  authorName - Author name(s), comma separated', body_font),
    ('  sourceUrl - DOI link or source URL', body_font),
    ('  imageUrl - Direct URL to article image (leave blank for auto placeholder)', body_font),
    ('  isPublished - TRUE to publish immediately, FALSE to save as draft', body_font),
    ('', None),
    ('DATA SOURCE:', section_font),
    ('  All articles sourced from PubMed (pubmed.ncbi.nlm.nih.gov)', note_font),
    ('  DOI links provided for proper attribution to original authors', note_font),
]

for i, (text, font) in enumerate(instr_rows, 1):
    cell = ins.cell(row=i, column=1, value=text)
    if font:
        cell.font = font

wb.save('article-upload-template.xlsx')
print('Done: 13 PubMed articles across 4 sub-categories')
